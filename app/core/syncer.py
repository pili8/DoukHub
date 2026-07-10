"""飞书同步引擎 — 读取采集表 → 解析短链接 → 获取账号信息 → 写入账号表"""
import asyncio
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from .collector import Account, Collector
from .feishu import FeishuClient
from .link_resolver import detect_platform, extract_sec_user_id, extract_url_from_text

logger = logging.getLogger("doukhub.syncer")


# 飞书账号表字段名 → Account 属性名 的映射
ACCOUNT_FIELD_MAP = {
    "账号名称": "name",
    "平台": "platform",
    "链接": "link",
    "采集类型": "collection_type",
    "代理": "proxy",
    "启用": "enabled",
    "等级": "rating",
    "标签": "tags",
    "备注": "note",
    "sec_user_id": "sec_user_id",
    "昵称": "nickname",
    "粉丝数": "follower_count",
    "作品数": "aweme_count",
    "签名": "signature",
    "头像": "avatar",
    "同步时间": "synced_at",
    "已获取信息": "info_fetched",
}

# 采集表字段名（只读取必要信息，不反写）
COLLECTION_FIELDS = {
    "地址": "link",           # 必填：短链接
    "等级": "rating",         # 必填：评级
    "标签": "tags",           # 可选：标签
    "账号名称": "name",       # 可选：账号名称
    "平台": "platform",       # 可选：平台（可自动识别）
    "备注": "note",           # 可选：备注
    # 以下字段如果采集表有则读取，没有则忽略
    "sec_user_id": "sec_user_id",
    "粉丝数": "follower_count",
    "作品数": "aweme_count",
    "同步状态": "sync_status",
    "同步时间": "synced_at",
}

# Cookie 表字段名
COOKIE_FIELDS = {
    "Cookie": "cookie",
    "平台": "platform",
    "状态": "status",
    "启用": "enabled",
    "备注": "note",
    "最后验证时间": "verified_at",
}


def _parse_rating(value: Any) -> int:
    """解析评级字段 — 支持数字和文本混合格式（如 '个3' → 3）"""
    if isinstance(value, (int, float)):
        return max(1, min(4, int(value)))
    if isinstance(value, str):
        # 提取文本中的数字
        import re
        numbers = re.findall(r"\d+", value)
        if numbers:
            return max(1, min(4, int(numbers[0])))
    return 3  # 默认 3 星


def _parse_tags(value: Any) -> list[str]:
    """解析标签字段"""
    if isinstance(value, list):
        return [str(v) for v in value]
    if isinstance(value, str):
        return [t.strip() for t in value.split(",") if t.strip()]
    return []


def _parse_enabled(value: Any) -> bool:
    """解析启用字段"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("true", "是", "1", "yes")
    if isinstance(value, (int, float)):
        return bool(value)
    return True


def _parse_collection_record(record: dict) -> dict:
    """解析采集表记录为 dict"""
    fields = record.get("fields", {})
    data = {"record_id": record.get("record_id", "")}

    for feishu_name, attr_name in COLLECTION_FIELDS.items():
        value = fields.get(feishu_name)
        if value is None:
            continue

        if attr_name == "rating":
            data[attr_name] = _parse_rating(value)
        elif attr_name == "tags":
            data[attr_name] = _parse_tags(value)
        elif attr_name == "enabled":
            data[attr_name] = _parse_enabled(value)
        elif attr_name in ("follower_count", "aweme_count"):
            try:
                data[attr_name] = int(value)
            except (ValueError, TypeError):
                data[attr_name] = 0
        elif attr_name == "link":
            if isinstance(value, dict):
                data[attr_name] = value.get("link", value.get("text", ""))
            else:
                data[attr_name] = str(value)
        else:
            data[attr_name] = str(value) if value else ""

    return data


def _parse_cookie_record(record: dict) -> dict:
    """解析 Cookie 表记录"""
    fields = record.get("fields", {})
    data = {"record_id": record.get("record_id", "")}

    for feishu_name, attr_name in COOKIE_FIELDS.items():
        value = fields.get(feishu_name)
        if value is None:
            continue
        if attr_name == "enabled":
            data[attr_name] = _parse_enabled(value)
        else:
            data[attr_name] = str(value) if value else ""

    return data


class SyncResult:
    """同步结果"""

    def __init__(self):
        self.total: int = 0
        self.new_accounts: int = 0
        self.updated_accounts: int = 0
        self.errors: list[str] = []
        self.success: bool = False
        self.message: str = ""
        self.api_calls: int = 0  # API 调用次数

    @property
    def summary(self) -> str:
        parts = [f"共 {self.total} 条记录"]
        if self.new_accounts:
            parts.append(f"新增 {self.new_accounts} 个")
        if self.updated_accounts:
            parts.append(f"更新 {self.updated_accounts} 个")
        if self.errors:
            parts.append(f"{len(self.errors)} 个错误")
        if self.api_calls:
            parts.append(f"{self.api_calls} 次 API 调用")
        return "，".join(parts)


class Syncer:
    """飞书同步引擎"""

    def __init__(
        self,
        feishu: FeishuClient,
        collector: Collector,
        app_token: str,
        collection_table_id: str,
        account_table_id: str,
        cookie_table_id: str = "",
        data_dir: Path = None,
    ):
        self.feishu = feishu
        self.collector = collector
        self.app_token = app_token
        self.collection_table_id = collection_table_id
        self.account_table_id = account_table_id
        self.cookie_table_id = cookie_table_id
        self.data_dir = data_dir
        self.accounts_file = data_dir / "accounts.xlsx" if data_dir else None
        self.cookies_file = data_dir / "cookies.xlsx" if data_dir else None

    async def sync(self) -> SyncResult:
        """执行同步：采集表 → 解析短链接 → 写入账号表（第一阶段：快速同步）"""
        result = SyncResult()

        try:
            # 1. 读取采集表全部记录
            logger.info("正在连接飞书...")
            records = self.feishu.get_all_records(self.app_token, self.collection_table_id)
            result.total = len(records)
            logger.info(f"读取采集表完成: {result.total} 条记录")

            # 2. 解析记录
            entries = [_parse_collection_record(r) for r in records]
            entries = [e for e in entries if e.get("link")]  # 过滤空链接

            # 3. 获取 Cookie 池
            cookies = self.load_local_cookies()
            active_cookies = [c["cookie"] for c in cookies if c.get("enabled", True) and c.get("status", "正常") == "正常"]

            # 4. 加载已有账号（按 sec_user_id 去重）
            existing_accounts = {}
            if self.accounts_file and self.accounts_file.exists():
                for acc in self.load_local_accounts():
                    if acc.sec_user_id:
                        existing_accounts[acc.sec_user_id] = acc

            # 5. 逐条处理（第一阶段：只解析短链接，不获取账号信息）
            for i, entry in enumerate(entries):
                link = entry["link"]
                rating = entry.get("rating", 3)
                record_id = entry["record_id"]
                # 检查采集表是否已有 sec_user_id
                existing_sec_user_id = entry.get("sec_user_id", "")

                logger.info(f"处理 [{i+1}/{len(entries)}]: {link}")

                try:
                    # 补全短链接前缀（如果只有缩略路径如 m3HL2u1R1YM）
                    if link and not link.startswith("http"):
                        link = f"https://v.douyin.com/{link}"

                    # 平台识别（纯正则）
                    platform = entry.get("platform", "") or detect_platform(link)

                    # 如果采集表已有 sec_user_id，跳过短链接解析
                    if existing_sec_user_id:
                        sec_user_id = existing_sec_user_id
                        resolved_url = link
                        logger.info(f"  使用采集表已有 sec_user_id: {sec_user_id}（跳过 API 调用）")
                    else:
                        # 通过 TTD API 解析短链接 → 获取完整 URL
                        cookie = active_cookies[0] if active_cookies else ""
                        resolved_url = await self.collector.resolve_short_url(link, platform)
                        result.api_calls += 1  # 记录 API 调用
                        logger.info(f"  短链接解析: {resolved_url}")

                        # 从完整 URL 中提取 sec_user_id（纯正则）
                        sec_user_id = extract_sec_user_id(resolved_url, platform)

                        if not sec_user_id:
                            self._update_collection_status(record_id, "失败", "无法解析短链接")
                            result.errors.append(f"{link}: 无法解析")
                            continue

                    # 检查是否已有账号
                    existing_account = existing_accounts.get(sec_user_id)

                    if existing_account:
                        # 更新已有账号
                        account = existing_account
                        account.link = resolved_url or link
                        account.rating = rating
                        account.tags = entry.get("tags", [])
                        account.note = entry.get("note", "")
                        account.synced_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        result.updated_accounts += 1
                        logger.info(f"  更新账号: {account.name or sec_user_id}")
                    else:
                        # 创建新账号（不获取详细信息）
                        account = Account(
                            name=entry.get("name", ""),
                            platform=platform,
                            link=resolved_url or link,
                            rating=rating,
                            tags=entry.get("tags", []),
                            note=entry.get("note", ""),
                            sec_user_id=sec_user_id,
                            synced_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            enabled=True,
                            info_fetched=False,  # 标记为未获取信息
                        )
                        result.new_accounts += 1
                        existing_accounts[sec_user_id] = account
                        logger.info(f"  新增账号: {account.name or sec_user_id}")

                    # 更新采集表状态（包含 sec_user_id）
                    self._update_collection_status(record_id, "已同步", "", sec_user_id)

                except Exception as e:
                    self._update_collection_status(record_id, "失败", str(e))
                    result.errors.append(f"{link}: {e}")

            # 6. 保存本地账号缓存
            logger.info("保存本地账号缓存...")
            self._save_local_xlsx(list(existing_accounts.values()))

            # 7. 写入飞书账号表
            if self.account_table_id:
                logger.info("写入飞书账号表...")
                self._sync_to_feishu_account_table(list(existing_accounts.values()), result)

            result.success = True
            result.message = result.summary
            logger.info(f"同步完成: {result.summary}")

        except Exception as e:
            result.success = False
            result.message = f"同步失败: {e}"
            logger.error(f"同步失败: {e}")

        return result

    async def fetch_account_info(self) -> SyncResult:
        """获取账号详细信息（第二阶段：异步获取）"""
        result = SyncResult()

        try:
            # 1. 加载本地账号
            if not self.accounts_file or not self.accounts_file.exists():
                result.success = False
                result.message = "没有本地账号缓存，请先执行同步"
                return result

            accounts = self.load_local_accounts()
            result.total = len(accounts)

            # 2. 筛选未获取信息的账号
            accounts_to_fetch = [acc for acc in accounts if not acc.info_fetched and acc.sec_user_id]
            logger.info(f"需要获取信息的账号: {len(accounts_to_fetch)} 个")

            if not accounts_to_fetch:
                result.success = True
                result.message = "所有账号信息已获取"
                return result

            # 3. 获取 Cookie 池
            cookies = self.load_local_cookies()
            active_cookies = [c["cookie"] for c in cookies if c.get("enabled", True) and c.get("status", "正常") == "正常"]

            # 4. 逐条获取账号信息
            for i, account in enumerate(accounts_to_fetch):
                logger.info(f"获取账号信息 [{i+1}/{len(accounts_to_fetch)}]: {account.sec_user_id}")

                try:
                    cookie = active_cookies[0] if active_cookies else ""
                    info = await self.collector.get_account_info(account.sec_user_id, account.platform, cookie)
                    result.api_calls += 1

                    # 更新账号信息
                    if info.get("nickname"):
                        account.nickname = info.get("nickname", "")
                        account.follower_count = info.get("follower_count", 0)
                        account.aweme_count = info.get("aweme_count", 0)
                        account.signature = info.get("signature", "")
                        account.avatar = info.get("avatar", "")
                        account.name = account.name or account.nickname
                        account.info_fetched = True
                        account.synced_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        result.updated_accounts += 1
                        logger.info(f"  获取成功: {account.nickname} ({account.follower_count} 粉丝)")
                    else:
                        result.errors.append(f"{account.sec_user_id}: 获取信息失败")
                        logger.warning(f"  获取失败: {account.sec_user_id}")

                except Exception as e:
                    result.errors.append(f"{account.sec_user_id}: {e}")
                    logger.error(f"  获取异常: {account.sec_user_id} - {e}")

            # 5. 保存本地账号缓存
            logger.info("保存本地账号缓存...")
            self._save_local_xlsx(accounts)

            # 6. 写入飞书账号表
            if self.account_table_id:
                logger.info("写入飞书账号表...")
                self._sync_to_feishu_account_table(accounts, result)

            result.success = True
            result.message = f"获取完成: {result.updated_accounts} 个账号"
            logger.info(f"获取完成: {result.message}")

        except Exception as e:
            result.success = False
            result.message = f"获取失败: {e}"
            logger.error(f"获取失败: {e}")

        return result

    def _update_collection_status(self, record_id: str, status: str, error: str = "", sec_user_id: str = "") -> None:
        """更新采集表中的同步状态"""
        if not record_id:
            return
        fields = {
            "同步状态": status,
            "同步时间": int(datetime.now().timestamp() * 1000),  # 飞书日期字段需要毫秒时间戳
        }
        if sec_user_id:
            fields["sec_user_id"] = sec_user_id
        try:
            self.feishu.update_record(
                self.app_token, self.collection_table_id, record_id, fields
            )
        except Exception:
            pass  # 回填失败不影响主流程

    def _save_local_xlsx(self, accounts: list[Account]) -> None:
        """保存账号列表到本地 XLSX"""
        if not self.accounts_file:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "账号列表"

        headers = list(ACCOUNT_FIELD_MAP.keys())
        ws.append(headers)

        for acc in accounts:
            row = []
            for feishu_name, attr_name in ACCOUNT_FIELD_MAP.items():
                value = getattr(acc, attr_name, "")
                if isinstance(value, list):
                    value = ", ".join(value)
                row.append(value)
            ws.append(row)

        self.accounts_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self.accounts_file))

    def _sync_to_feishu_account_table(self, accounts: list[Account], result: SyncResult) -> None:
        """将账号数据同步到飞书账号表"""
        if not self.account_table_id:
            return

        try:
            # 读取飞书账号表现有记录（按 sec_user_id 索引）
            existing_records = self.feishu.get_all_records(self.app_token, self.account_table_id)
            existing_by_sec_id = {}
            for record in existing_records:
                fields = record.get("fields", {})
                sec_id = fields.get("sec_user_id", "")
                if sec_id:
                    existing_by_sec_id[sec_id] = record.get("record_id", "")

            # 准备要写入的记录
            records_to_create = []
            records_to_update = []

            for acc in accounts:
                if not acc.sec_user_id:
                    continue

                # 构建字段（反向映射：attr_name → feishu_name）
                fields = {}
                for feishu_name, attr_name in ACCOUNT_FIELD_MAP.items():
                    value = getattr(acc, attr_name, "")
                    if isinstance(value, list):
                        value = value  # 多选字段保持列表
                    elif isinstance(value, bool):
                        value = value
                    elif value != "" and value is not None:
                        pass
                    else:
                        continue
                    fields[feishu_name] = value

                if acc.sec_user_id in existing_by_sec_id:
                    # 更新现有记录
                    records_to_update.append({
                        "record_id": existing_by_sec_id[acc.sec_user_id],
                        "fields": fields,
                    })
                else:
                    # 创建新记录
                    records_to_create.append({"fields": fields})

            # 批量创建
            if records_to_create:
                batch_size = 500
                for i in range(0, len(records_to_create), batch_size):
                    batch = records_to_create[i:i + batch_size]
                    self.feishu.batch_create_records(self.app_token, self.account_table_id, batch)
                    result.api_calls += 1

            # 批量更新
            if records_to_update:
                batch_size = 500
                for i in range(0, len(records_to_update), batch_size):
                    batch = records_to_update[i:i + batch_size]
                    self.feishu.batch_update_records(self.app_token, self.account_table_id, batch)
                    result.api_calls += 1

        except Exception as e:
            result.errors.append(f"写入账号表失败: {e}")

    def load_local_accounts(self) -> list[Account]:
        """从本地 XLSX 加载账号列表"""
        if not self.accounts_file or not self.accounts_file.exists():
            return []

        wb = openpyxl.load_workbook(str(self.accounts_file))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = list(ACCOUNT_FIELD_MAP.keys())

        accounts = []
        for row in rows:
            data = dict(zip(headers, row))
            account = Account()
            for feishu_name, attr_name in ACCOUNT_FIELD_MAP.items():
                value = data.get(feishu_name)
                if attr_name == "rating":
                    setattr(account, attr_name, _parse_rating(value))
                elif attr_name == "tags":
                    setattr(account, attr_name, _parse_tags(value))
                elif attr_name == "enabled":
                    setattr(account, attr_name, _parse_enabled(value))
                elif attr_name in ("follower_count", "aweme_count"):
                    try:
                        setattr(account, attr_name, int(value or 0))
                    except (ValueError, TypeError):
                        setattr(account, attr_name, 0)
                else:
                    setattr(account, attr_name, str(value) if value else "")
            accounts.append(account)

        wb.close()
        return accounts

    def load_local_cookies(self) -> list[dict]:
        """从本地 XLSX 加载 Cookie 列表"""
        if not self.cookies_file or not self.cookies_file.exists():
            return []

        wb = openpyxl.load_workbook(str(self.cookies_file))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = list(COOKIE_FIELDS.keys())

        cookies = []
        for row in rows:
            data = dict(zip(headers, row))
            cookie = {}
            for feishu_name, attr_name in COOKIE_FIELDS.items():
                value = data.get(feishu_name)
                if attr_name == "enabled":
                    cookie[attr_name] = _parse_enabled(value)
                else:
                    cookie[attr_name] = str(value) if value else ""
            cookies.append(cookie)

        wb.close()
        return cookies

    def sync_cookies(self) -> dict:
        """同步 Cookie 表到本地"""
        if not self.cookie_table_id:
            return {"success": False, "message": "未配置 Cookie 表"}

        try:
            records = self.feishu.get_all_records(self.app_token, self.cookie_table_id)
            cookies = [_parse_cookie_record(r) for r in records]

            # 保存到本地
            if self.cookies_file:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Cookie 池"
                headers = list(COOKIE_FIELDS.keys())
                ws.append(headers)
                for c in cookies:
                    row = [c.get(attr, "") for attr in COOKIE_FIELDS.values()]
                    ws.append(row)
                self.cookies_file.parent.mkdir(parents=True, exist_ok=True)
                wb.save(str(self.cookies_file))

            return {"success": True, "message": f"同步 {len(cookies)} 个 Cookie"}
        except Exception as e:
            return {"success": False, "message": f"同步失败: {e}"}
